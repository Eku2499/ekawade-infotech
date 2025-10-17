import openpyxl
from flask import Flask, request, jsonify
import os

app = Flask(__name__)
EXCEL_FILE = 'booking.xlsx'

# Ensure Excel file exists with headers
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Name', 'Email', 'Date of Appointment', 'Phone Number'])
        wb.save(EXCEL_FILE)

@app.route('/book', methods=['POST'])
def book():
    data = request.json
    name = data.get('name')
    email = data.get('email')
    date = data.get('date')
    phone = data.get('phone')
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, email, date, phone])
    wb.save(EXCEL_FILE)
    return jsonify({'status': 'success'})

@app.route('/bookings', methods=['GET'])
def get_bookings():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    bookings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        bookings.append({
            'name': row[0],
            'email': row[1],
            'date': row[2],
            'phone': row[3]
        })
    return jsonify(bookings)

if __name__ == '__main__':
    init_excel()
    app.run(debug=True)
